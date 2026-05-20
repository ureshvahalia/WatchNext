"""FastAPI local server for the WatchNext rating UI."""
import os
import sys
import threading
import time
import webbrowser
from pathlib import Path
from typing import Optional

import openpyxl
import requests
import uvicorn
from fastapi import FastAPI
from fastapi.responses import FileResponse
from openpyxl.styles import Font
from pydantic import BaseModel

# ── Paths ─────────────────────────────────────────────────────────────────────
_UI_DIR = Path(__file__).parent
_ROOT   = _UI_DIR.parent
sys.path.insert(0, str(_ROOT / "recommender"))

from _common import (
    GENRE_MAP, MATCHES_FILE, RECS_FILE, WATCH_FILE,
    get_api_key, load_matches, load_quality_cache, load_recs,
    score_method_g,
)

TMDB_BASE    = "https://api.themoviedb.org/3"
TMDB_IMG     = "https://image.tmdb.org/t/p/w342"
COUNTRY      = os.environ.get("WATCHNEXT_COUNTRY", "IN")
POSTER_CACHE = _ROOT / "recommender" / "cache" / "poster_cache.json"
INDEX_HTML   = _UI_DIR / "index.html"

app = FastAPI(title="WatchNext")

# ── Shared state ──────────────────────────────────────────────────────────────
_file_lock   = threading.Lock()
_poster_lock = threading.Lock()
_poster_data: dict = {}

# Session caches — populated once, read many times
_matches_cache:   Optional[dict] = None
_recs_cache:      Optional[dict] = None
_watched_ids:     Optional[set]  = None
_quality_cache:   Optional[dict] = None


# ── Session cache helpers ─────────────────────────────────────────────────────

def _get_matches() -> dict:
    global _matches_cache
    if _matches_cache is None:
        _matches_cache = load_matches()
    return _matches_cache


def _get_recs() -> dict:
    global _recs_cache
    if _recs_cache is None:
        _recs_cache = load_recs()
    return _recs_cache


def _get_watched_ids() -> set:
    global _watched_ids
    if _watched_ids is None:
        _watched_ids = {str(m["tmdb_id"]) for m in _get_matches().values() if m.get("matched")}
    return _watched_ids


def _get_quality_cache() -> dict:
    global _quality_cache
    if _quality_cache is None:
        _quality_cache = load_quality_cache()
    return _quality_cache


# ── Rating / scoring ──────────────────────────────────────────────────────────

def _load_ratings() -> dict:
    """Return {title: float_or_None} for all matched watched titles."""
    if not WATCH_FILE.exists():
        return {}
    with _file_lock:
        wb = openpyxl.load_workbook(WATCH_FILE)
    ws      = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col     = {h: i for i, h in enumerate(headers)}
    name_idx   = col.get("Name", 1)
    rating_idx = col.get("Your Rating")

    result: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        if not name:
            continue
        rating = None
        if rating_idx is not None and rating_idx < len(row):
            r = row[rating_idx]
            if r is not None:
                try:
                    val = float(r)
                    if 1.0 <= val <= 5.0:
                        rating = val
                except (ValueError, TypeError):
                    pass
        result[str(name)] = rating
    return result


def _compute_recs(ratings: dict) -> list:
    """Method G — Hybrid 80/20 × Quality^1.0. Returns top-50 for the UI."""
    active = {k: v for k, v in ratings.items() if v is not None}
    ranked = score_method_g(
        _get_matches(), _get_recs(), active,
        _get_watched_ids(), _get_quality_cache(),
    )
    out = []
    for rank, item in enumerate(ranked[:50], 1):
        genres = [GENRE_MAP.get(gid, str(gid)) for gid in item["genre_ids"]]
        seg    = "movie" if item["type"] == "Movie" else "tv"
        out.append({
            "rank":     rank,
            "tmdb_id":  item["tmdb_id"],
            "type":     item["type"],
            "title":    item["title"],
            "year":     item["year"],
            "freq":     item["freq"],
            "score":    item["score"],
            "vote_avg": round(item["vote_average"], 1),
            "genres":   genres,
            "tmdb_url": f"https://www.themoviedb.org/{seg}/{item['tmdb_id']}",
        })
    return out


# ── Poster / enrichment ───────────────────────────────────────────────────────

def _fetch_enrichment(tmdb_id: int, media_type: str, api_key: str) -> dict:
    seg    = "movie" if media_type == "Movie" else "tv"
    result = {"poster_url": None, "overview": "", "watch_url": None}
    if not api_key:
        return result

    if api_key.startswith("eyJ"):
        headers = {"Authorization": f"Bearer {api_key}"}
        params  = {}
    else:
        headers = {}
        params  = {"api_key": api_key}

    try:
        r = requests.get(f"{TMDB_BASE}/{seg}/{tmdb_id}",
                         headers=headers, params=params, timeout=8)
        if r.ok:
            d  = r.json()
            pp = d.get("poster_path")
            result["poster_url"] = f"{TMDB_IMG}{pp}" if pp else None
            result["overview"]   = d.get("overview", "")
        time.sleep(0.1)

        r2 = requests.get(f"{TMDB_BASE}/{seg}/{tmdb_id}/watch/providers",
                          headers=headers, params=params, timeout=8)
        if r2.ok:
            providers = r2.json().get("results", {})
            for country in [COUNTRY, "US", "GB"]:
                if country in providers and providers[country].get("link"):
                    result["watch_url"] = providers[country]["link"]
                    break
        time.sleep(0.1)
    except Exception:
        pass

    return result


# ── Startup ───────────────────────────────────────────────────────────────────

@app.on_event("startup")
def _startup():
    import json
    global _poster_data
    if POSTER_CACHE.exists():
        try:
            _poster_data = json.loads(POSTER_CACHE.read_text(encoding="utf-8"))
        except Exception:
            pass


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/")
def index():
    return FileResponse(str(INDEX_HTML))


@app.get("/api/watched")
def api_watched():
    """All matched watched titles with their current rating (or null)."""
    if not WATCH_FILE.exists() or not MATCHES_FILE.exists():
        return []
    matches = _get_matches()
    ratings = _load_ratings()

    items = []
    for name, rating in ratings.items():
        m = matches.get(name, {})
        if not m.get("matched"):
            continue
        items.append({
            "name":    name,
            "type":    m.get("media_type", ""),
            "year":    m.get("year", ""),
            "tmdb_id": m["tmdb_id"],
            "rating":  rating,
        })

    # Unrated first (then alphabetical within each group)
    items.sort(key=lambda x: (x["rating"] is not None, x["name"].lower()))
    return items


@app.get("/api/recommendations")
def api_recommendations():
    """Current recommendations ranked by weighted score."""
    if not WATCH_FILE.exists() or not MATCHES_FILE.exists() or not RECS_FILE.exists():
        return []
    ratings = {k: v for k, v in _load_ratings().items() if v is not None}
    return _compute_recs(ratings)


class RateBody(BaseModel):
    name:   str
    rating: Optional[float] = None  # None clears the rating


@app.post("/api/rate")
def api_rate(body: RateBody):
    """Persist a rating to watch_history.xlsx and return updated recommendations."""
    if not WATCH_FILE.exists():
        return []

    with _file_lock:
        wb      = openpyxl.load_workbook(WATCH_FILE)
        ws      = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col     = {h: i for i, h in enumerate(headers)}
        name_idx    = col.get("Name", 1)
        rating_col0 = col.get("Your Rating")        # 0-based

        if rating_col0 is None:
            rating_col1 = ws.max_column + 1         # 1-based, new column
            cell       = ws.cell(1, rating_col1)
            cell.value = "Your Rating"
            cell.font  = Font(bold=True)
        else:
            rating_col1 = rating_col0 + 1           # 0-based -> 1-based

        for row_num in range(2, ws.max_row + 1):
            if ws.cell(row_num, name_idx + 1).value == body.name:
                ws.cell(row_num, rating_col1).value = body.rating
                break

        try:
            wb.save(WATCH_FILE)
        except PermissionError:
            return {"error": "watch_history.xlsx is open in another program"}

        # Re-read all ratings from the just-saved workbook
        ratings: dict = {}
        for row_num in range(2, ws.max_row + 1):
            n = ws.cell(row_num, name_idx + 1).value
            r = ws.cell(row_num, rating_col1).value
            if n and r is not None:
                try:
                    val = float(r)
                    if 1.0 <= val <= 5.0:
                        ratings[str(n)] = val
                except (ValueError, TypeError):
                    pass

    return _compute_recs(ratings)


class EnrichBody(BaseModel):
    items: list[dict]   # [{"tmdb_id": int, "type": "Movie"|"Series"}, ...]


@app.post("/api/enrich")
def api_enrich(body: EnrichBody):
    """Return poster URLs, overviews, and watch links for a batch of TMDB IDs.

    Results are cached in poster_cache.json so repeated requests are instant.
    """
    api_key = get_api_key()
    result  = {}

    for item in body.items:
        tid = str(item.get("tmdb_id", ""))
        if not tid:
            continue

        with _poster_lock:
            if tid in _poster_data:
                result[tid] = _poster_data[tid]
                continue

        data = _fetch_enrichment(int(tid), item.get("type", "Movie"), api_key)

        with _poster_lock:
            _poster_data[tid] = data
        result[tid] = data

    # Persist cache (best-effort; failure just means next run refetches)
    try:
        import json
        POSTER_CACHE.parent.mkdir(parents=True, exist_ok=True)
        with _poster_lock:
            POSTER_CACHE.write_text(
                json.dumps(_poster_data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    except Exception:
        pass

    return result


# ── Entry point ───────────────────────────────────────────────────────────────

def launch(host: str = "127.0.0.1", port: int = 8765):
    url = f"http://{host}:{port}"
    print(f"\nWatchNext UI  →  {url}")
    print("Press Ctrl+C to stop.\n")
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    uvicorn.run(app, host=host, port=port, log_level="warning")


if __name__ == "__main__":
    launch()
