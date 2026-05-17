"""
Step 6 -- Aggregate platform recommendation carousels into a ranked list.

Reads recommender/cache/platform_recs.json and counts how many watched titles
recommended each candidate, then ranks by that frequency score.

Already-watched titles are filtered out by URL (exact) and title (fuzzy >= 90).

Where a recommended title's URL is also in your watch history, the TMDB
metadata (rating, genres, year) is pulled from the existing match cache to
enrich the output.  Otherwise the row shows platform data only.

Outputs output/platform_recommendations.xlsx with:
  Recommendations -- ranked list
  Audit           -- one row per (watched title -> recommendation) pair
  No_Recs         -- watched titles whose page returned zero recommendations

Usage:
  python recommender/06_aggregate_platform.py
"""
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from thefuzz import fuzz

sys.path.insert(0, str(Path(__file__).parent))
from _common import (
    WATCH_FILE, MATCHES_FILE, GENRE_MAP, TMDB_BASE_URL,
    load_matches, save_json,
)

_HOME          = os.environ.get('WATCHNEXT_HOME')
_REC_DIR       = (Path(_HOME) / "recommender") if _HOME else Path(__file__).parent
PLATFORM_CACHE = _REC_DIR / "cache" / "platform_recs.json"
OUT_FILE       = (Path(_HOME) / "output" / "platform_recommendations.xlsx") if _HOME else Path(__file__).parent.parent / "output" / "platform_recommendations.xlsx"
FILTER_THRESHOLD = 90
MAX_SOURCES_INLINE = 20


# ── helpers ────────────────────────────────────────────────────────────────────

def load_platform_cache() -> dict:
    if not PLATFORM_CACHE.exists():
        print("No platform recs cache found. Run  run.bat ScrapePlatform  first.")
        sys.exit(1)
    return json.loads(PLATFORM_CACHE.read_text(encoding="utf-8"))


def load_watched_urls() -> set:
    """Return set of canonical watched URLs (query strings stripped)."""
    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers    = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col        = {h: i for i, h in enumerate(headers)}
    url_1based = col.get("url", 3) + 1
    urls = set()
    for r_idx in range(2, ws.max_row + 1):
        h = ws.cell(r_idx, url_1based).hyperlink
        if h:
            url = getattr(h, "target", "") or str(h)
            urls.add(url.rstrip("/"))
    return urls


def load_watched_titles() -> set:
    """Return set of lowercased watched title names."""
    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers  = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col      = {h: i for i, h in enumerate(headers)}
    name_idx = col.get("Name", 1)
    return {
        str(row[name_idx]).lower().strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[name_idx]
    }


def is_watched(url: str, title: str, watched_urls: set, watched_titles: set) -> bool:
    if url.rstrip("/") in watched_urls:
        return True
    t = title.lower().strip()
    if t in watched_titles:
        return True
    return any(fuzz.token_sort_ratio(t, wt) >= FILTER_THRESHOLD for wt in watched_titles)


def tmdb_meta_for_url(url: str, title: str, matches: dict) -> dict | None:
    """Try to find TMDB metadata for a recommended title via its URL or name."""
    # Check if the recommended URL matches any watched title's stored URL
    # (won't find new titles, but enriches anything already in our watch history)
    clean_title = title.lower().strip()
    for name, m in matches.items():
        if not m.get("matched"):
            continue
        if fuzz.token_sort_ratio(clean_title, name.lower().strip()) >= 90:
            return m
    return None


# ── main ───────────────────────────────────────────────────────────────────────

def main():
    platform_cache = load_platform_cache()
    matches        = load_matches()
    watched_urls   = load_watched_urls()
    watched_titles = load_watched_titles()

    # scores[canonical_url] = {count, title, platform, url, sources[], tmdb}
    scores: dict[str, dict] = defaultdict(lambda: {
        "count": 0, "title": "", "platform": "", "url": "",
        "sources": [], "tmdb": None,
    })
    audit_rows:  list[tuple] = []
    no_recs:     list[tuple] = []   # (source_title, platform)

    print(f"Aggregating platform recommendations from {len(platform_cache)} scraped titles...")

    for source_url, entry in platform_cache.items():
        source_name = entry.get("source_title", source_url)
        platform    = entry.get("platform", "")
        recs        = entry.get("recs", [])

        if not recs:
            no_recs.append((source_name, platform))
            continue

        seen_this_source: set[str] = set()
        for rec in recs:
            rec_url   = rec.get("url", "").rstrip("/")
            rec_title = rec.get("title", "").strip()
            if not rec_url or not rec_title:
                continue
            if rec_url in seen_this_source:
                continue
            seen_this_source.add(rec_url)

            if is_watched(rec_url, rec_title, watched_urls, watched_titles):
                continue

            s = scores[rec_url]
            s["count"]    += 1
            s["title"]     = rec_title      # last-seen title text for this URL
            s["platform"]  = platform
            s["url"]       = rec_url
            s["sources"].append(source_name)
            if s["tmdb"] is None:
                s["tmdb"] = tmdb_meta_for_url(rec_url, rec_title, matches)

            audit_rows.append((
                source_name, platform,
                rec_title, rec_url,
            ))

    ranked = sorted(
        [(url, v) for url, v in scores.items() if v["title"]],
        key=lambda x: -x[1]["count"],
    )
    rank_of = {url: r for r, (url, _) in enumerate(ranked, 1)}

    print(f"Found {len(ranked)} unique candidates.")

    wb = openpyxl.Workbook()

    # ── Recommendations sheet ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Recommendations"

    headers = [
        "Rank", "Title", "Platform", "Score",
        "TMDB Rating", "Genres", "Year",       # populated where known
        "Platform Link", "Recommended By",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rank, (url, v) in enumerate(ranked, 1):
        tmdb = v["tmdb"] or {}
        genre_ids = tmdb.get("genre_ids", [])
        genres    = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in genre_ids)
        sources   = v["sources"]
        if len(sources) <= MAX_SOURCES_INLINE:
            sources_str = " | ".join(sources)
        else:
            sources_str = " | ".join(sources[:MAX_SOURCES_INLINE]) + f"  … (+{len(sources)-MAX_SOURCES_INLINE} more)"

        ws.append([
            rank, v["title"], v["platform"], v["count"],
            round(tmdb.get("vote_average") or 0, 1) or "",
            genres or "",
            tmdb.get("year", "") or "",
            url, sources_str,
        ])

    # ── Audit sheet ────────────────────────────────────────────────────────────
    ws_audit = wb.create_sheet("Audit")
    ws_audit.append(["Watched Title", "Platform", "Recommended Title", "Platform Link", "Rec Rank"])
    for cell in ws_audit[1]:
        cell.font = Font(bold=True)

    audit_rows.sort(key=lambda r: rank_of.get(r[3], 999999))
    for source_name, platform, rec_title, rec_url in audit_rows:
        ws_audit.append([source_name, platform, rec_title, rec_url, rank_of.get(rec_url, "")])

    # ── No_Recs sheet ──────────────────────────────────────────────────────────
    if no_recs:
        ws_nr = wb.create_sheet("No_Recs")
        ws_nr.append(["Watched Title", "Platform"])
        ws_nr[1][0].font = Font(bold=True)
        ws_nr[1][1].font = Font(bold=True)
        for name, plat in sorted(no_recs):
            ws_nr.append([name, plat])

    try:
        wb.save(OUT_FILE)
    except PermissionError:
        print(f"ERROR: Cannot write {OUT_FILE.name} — close it in Excel and retry.")
        sys.exit(1)

    print(f"Saved {len(ranked)} recommendations to {OUT_FILE}")
    print(f"Audit sheet: {len(audit_rows):,} source→recommendation rows")
    if no_recs:
        print(f"{len(no_recs)} titles returned no recs (see 'No_Recs' sheet)")


if __name__ == "__main__":
    main()
