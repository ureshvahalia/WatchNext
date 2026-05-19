"""
Step 3 -- Aggregate and rank recommendations.

Counts how many of your watched titles recommended each candidate, then ranks
by a combined score:

    score = freq * log(1 + tmdb_vote_average)

where freq is the raw count of source titles and log(1 + rating) blends in
TMDB quality without letting it dominate. Titles with no recorded TMDB rating
are treated as 5.0 (neutral midpoint) so they are not unfairly zeroed out.

Titles already in your watch history are filtered out (exact + fuzzy >= 90).

Outputs output/recommendations.xlsx with three sheets:
  Recommendations -- ranked list with a "Recommended By" column
  Audit           -- one row per (watched title -> recommendation) pair;
                     filter by column A to see everything a specific title surfaced,
                     or filter by column C to see all sources for one recommendation
  Unmatched       -- titles TMDB could not identify (shown only if non-empty)

Usage:
  python recommender/03_aggregate.py
"""
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))
from _common import (
    OUT_FILE, GENRE_MAP, TMDB_BASE_URL,
    load_matches, load_recs, watched_title_set, is_already_watched,
    tmdb_score_factor,
)

MAX_SOURCES_INLINE = 20   # max source titles listed in the "Recommended By" cell


def main():
    matches = load_matches()
    recs    = load_recs()
    watched = watched_title_set()

    if not matches:
        print("No TMDB matches found. Run  run.bat Match  first.")
        sys.exit(1)
    if not recs:
        print("No recommendation data found. Run  run.bat Recommend  first.")
        sys.exit(1)

    # Build lookup: TMDB source ID -> our watched title name
    id_to_name = {
        str(m["tmdb_id"]): name
        for name, m in matches.items()
        if m.get("matched")
    }

    watched_ids = {str(m["tmdb_id"]) for m in matches.values() if m.get("matched")}
    unmatched   = [
        (name, m.get("media_type", ""))
        for name, m in matches.items()
        if not m.get("matched")
    ]

    # scores[candidate_tid] = {count, data, sources: [watched_title_name, ...]}
    scores: dict[str, dict] = defaultdict(lambda: {"count": 0, "data": None, "sources": []})
    # audit_rows: flat list of (source_name, source_tmdb_id, rec_title, rec_type,
    #                           rec_year, rec_vote, rec_tid) for the Audit sheet
    audit_rows: list[tuple] = []

    print(f"Aggregating recommendations from {len(recs)} source titles...")
    for source_id, rec_list in recs.items():
        source_name = id_to_name.get(source_id, f"TMDB:{source_id}")
        seen_this_source: set[str] = set()
        for rec in rec_list:
            tid = str(rec["tmdb_id"])
            if tid in seen_this_source:
                continue
            seen_this_source.add(tid)
            if tid in watched_ids:
                continue
            if is_already_watched(rec["title"], watched):
                continue

            scores[tid]["count"] += 1
            scores[tid]["sources"].append(source_name)
            if scores[tid]["data"] is None:
                scores[tid]["data"] = rec

            audit_rows.append((
                source_name, source_id,
                rec["title"], rec.get("type", ""), rec.get("year", ""),
                rec.get("vote_average", 0), tid,
            ))

    ranked = sorted(
        [(tid, v) for tid, v in scores.items() if v["data"]],
        key=lambda x: -(x[1]["count"] * tmdb_score_factor(x[1]["data"].get("vote_average"))),
    )
    # Build rank lookup for the Audit sheet
    rank_of = {tid: rank for rank, (tid, _) in enumerate(ranked, 1)}

    print(f"Found {len(ranked)} unique candidates.")

    wb = openpyxl.Workbook()

    # ── Recommendations sheet ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Recommendations"

    headers = ["Rank", "Title", "Type", "Freq", "Score", "TMDB Rating",
               "Genres", "Year", "TMDB Link", "Recommended By"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rank, (tid, v) in enumerate(ranked, 1):
        d       = v["data"]
        genres  = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in d.get("genre_ids", []))
        mtype   = d.get("type", "")
        seg     = "movie" if mtype == "Movie" else "tv"
        link    = f"{TMDB_BASE_URL}/{seg}/{tid}"
        sources = v["sources"]
        if len(sources) <= MAX_SOURCES_INLINE:
            sources_str = " | ".join(sources)
        else:
            sources_str = " | ".join(sources[:MAX_SOURCES_INLINE]) + f"  … (+{len(sources) - MAX_SOURCES_INLINE} more)"
        vote_avg = d.get("vote_average") or 0
        score    = round(v["count"] * tmdb_score_factor(vote_avg), 2)
        ws.append([
            rank, d["title"], mtype, v["count"], score,
            round(vote_avg, 1),
            genres, d.get("year", ""), link, sources_str,
        ])

    # ── Audit sheet ────────────────────────────────────────────────────────────
    ws_audit = wb.create_sheet("Audit")
    audit_headers = [
        "Watched Title", "Watched TMDB ID",
        "Recommended Title", "Type", "Year", "TMDB Rating",
        "Recommended TMDB ID", "Rec Rank",
    ]
    ws_audit.append(audit_headers)
    for cell in ws_audit[1]:
        cell.font = Font(bold=True)

    # Sort audit rows by recommendation rank so the most-surfaced titles appear first
    audit_rows.sort(key=lambda r: rank_of.get(r[6], 999999))
    for source_name, source_id, rec_title, rec_type, rec_year, rec_vote, rec_tid in audit_rows:
        ws_audit.append([
            source_name, source_id,
            rec_title, rec_type, rec_year,
            round(rec_vote or 0, 1),
            rec_tid, rank_of.get(rec_tid, ""),
        ])

    # ── Unmatched sheet ────────────────────────────────────────────────────────
    if unmatched:
        ws_un = wb.create_sheet("Unmatched")
        ws_un.append(["Title", "Type"])
        ws_un[1][0].font = Font(bold=True)
        ws_un[1][1].font = Font(bold=True)
        for name, mtype in sorted(unmatched):
            ws_un.append([name, mtype])

    try:
        wb.save(OUT_FILE)
    except PermissionError:
        print(f"ERROR: Cannot write {OUT_FILE.name} -- close it in Excel and retry.")
        sys.exit(1)

    print(f"Saved {len(ranked)} recommendations to {OUT_FILE}")
    print(f"Audit sheet: {len(audit_rows):,} source→recommendation rows")
    if unmatched:
        print(f"{len(unmatched)} unmatched titles in 'Unmatched' sheet")


if __name__ == "__main__":
    main()
