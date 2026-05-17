"""
Step 1 -- Match watch_history.xlsx titles to TMDB IDs.

Matching strategy (in order of reliability):
  1. Fuzzy text search via TMDB /search/{movie|tv}
  2. For Netflix-sourced titles: confirm (or override) via TMDB /find/{netflix_id}
     - If both searches return the same TMDB ID  -> confirmed match
     - If they disagree                          -> Netflix ID wins; conflict logged
     - If fuzzy fails but Netflix ID succeeds    -> Netflix ID used; logged
     - If only fuzzy succeeds                    -> fuzzy result used

Conflict log: recommender/cache/match_conflicts.json
  Records every case where the two methods disagreed so you can audit them.

Other behaviour:
  - Adds a 'Your Rating' column to watch_history.xlsx if absent.
  - Caches results in recommender/cache/tmdb_matches.json; re-running skips
    already-cached titles unless --reset is passed.

Usage:
  python recommender/01_match_tmdb.py
  python recommender/01_match_tmdb.py --reset
"""
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from thefuzz import fuzz

sys.path.insert(0, str(Path(__file__).parent))
from _common import WATCH_FILE, MATCHES_FILE, require_api_key, load_matches, save_json, tmdb_get

TMDB_BASE          = "https://api.themoviedb.org/3"
MATCH_THRESHOLD    = 85   # strict threshold for normal cases
FALLBACK_THRESHOLD = 45   # used when the top result is clearly the lone best match

# Trailing parentheticals that platforms append but are not part of the title.
_PLATFORM_SUFFIX = re.compile(
    r'\s*\('
    r'(?:Hindi|Tamil|Telugu|Malayalam|Kannada|Marathi|Bengali|Punjabi|Gujarati|'
    r'English|French|Spanish|Portuguese|German|Italian|Korean|Japanese|Chinese|'
    r'Arabic|Turkish|Russian|Dutch|Swedish|Norwegian|Danish|Finnish|'
    r'(?:\w+\s+)?Dubbed|Version|Cut|Edition)'
    r'\)\s*$',
    re.IGNORECASE,
)
_REC_DIR        = (Path(os.environ['WATCHNEXT_HOME']) / "recommender") if 'WATCHNEXT_HOME' in os.environ else Path(__file__).parent
CONFLICTS_FILE  = _REC_DIR / "cache" / "match_conflicts.json"


# ── Title cleaning ────────────────────────────────────────────────────────────

def clean_for_search(title: str) -> str:
    """Strip platform-appended language/version suffixes before TMDB search.

    'Shabaash Mithu (Hindi)'  -> 'Shabaash Mithu'
    'RRR (Telugu)'            -> 'RRR'
    'Some Film (Hindi Dubbed)' -> 'Some Film'
    Year suffixes like '(2021)' and type tags like '(TV)' are left untouched.
    """
    return _PLATFORM_SUFFIX.sub("", title).strip()


# ── Netflix ID helpers ─────────────────────────────────────────────────────────

def extract_netflix_id(url: str) -> str | None:
    """Return the numeric Netflix title ID from a Netflix URL, or None."""
    if not url:
        return None
    m = re.search(r'netflix\.com/title/(\d+)', str(url))
    return m.group(1) if m else None


def find_by_netflix_id(netflix_id: str, api_key: str) -> dict | None:
    """Return the best TMDB result for a Netflix title ID, or None.

    Checks both movie_results and tv_results; returns whichever is non-empty
    (movies preferred when both are populated, which is rare).
    """
    try:
        r = tmdb_get(
            f"{TMDB_BASE}/find/{netflix_id}",
            api_key,
            params={"external_source": "netflix_id"},
        )
        data = r.json()
        for bucket, mtype in (("movie_results", "Movie"), ("tv_results", "Series")):
            results = data.get(bucket, [])
            if results:
                return {"result": results[0], "media_type": mtype}
    except Exception:
        pass
    return None


# ── Fuzzy search helpers ───────────────────────────────────────────────────────

def search_tmdb(title: str, media_type: str, api_key: str) -> list:
    endpoint = "movie" if media_type == "Movie" else "tv"
    r = tmdb_get(
        f"{TMDB_BASE}/search/{endpoint}",
        api_key,
        params={"query": title, "include_adult": False},
    )
    return r.json().get("results", [])


def best_match(title: str, results: list, media_type: str, watched_year: int | None = None):
    """Return (best_result, score) or (None, 0).

    Scoring strategy (max of three):
      token_sort_ratio  -- handles word-order differences
      partial_ratio     -- handles our title being a substring of TMDB's
                           (e.g. 'Borat Subsequent Moviefilm' vs the full long title)
      original_title    -- scores against the original-language title field
                           (catches films where TMDB's display title is a translation)

    Two acceptance thresholds:
      MATCH_THRESHOLD (85)   -- normal case
      FALLBACK_THRESHOLD (45) -- used when the top result is clearly the lone best
                                 candidate; trusts TMDB's own search ranking for
                                 foreign titles whose display title is a translation
                                 (e.g. 'Dolly Kitty Aur Woh Chamakte Sitare' returns
                                 the right film as the only result even though our
                                 romanized title doesn't fuzzy-match the English one)
    """
    title_key = "title"          if media_type == "Movie" else "name"
    orig_key  = "original_title" if media_type == "Movie" else "original_name"
    date_key  = "release_date"   if media_type == "Movie" else "first_air_date"

    candidates = []
    for r in results[:10]:
        tmdb_title = r.get(title_key, "") or ""
        orig_title = r.get(orig_key,  "") or ""
        score = max(
            fuzz.token_sort_ratio(title.lower(), tmdb_title.lower()),
            fuzz.partial_ratio(title.lower(),     tmdb_title.lower()),
            fuzz.token_sort_ratio(title.lower(), orig_title.lower()) if orig_title else 0,
        )
        candidates.append((score, r))
    candidates.sort(key=lambda x: -x[0])

    if not candidates:
        return None, 0

    top_score    = candidates[0][0]
    top_result   = candidates[0][1]
    second_score = candidates[1][0] if len(candidates) > 1 else 0

    def disambiguate_by_year(shortlist):
        """From a score-tied shortlist, prefer the result closest to watched_year."""
        if not watched_year or len(shortlist) < 2:
            return shortlist[0][1], shortlist[0][0]
        def year_dist(r):
            y = (r.get(date_key, "") or "")[:4]
            try:
                yr = int(y)
                return abs(yr - watched_year) if yr <= watched_year else 9999
            except ValueError:
                return 9999
        shortlist = sorted(shortlist, key=lambda x: year_dist(x[1]))
        return shortlist[0][1], shortlist[0][0]

    # ── strict path ───────────────────────────────────────────────────────────
    if top_score >= MATCH_THRESHOLD:
        close = [(s, r) for s, r in candidates if s >= top_score - 5]
        return disambiguate_by_year(close)

    # ── fallback path ─────────────────────────────────────────────────────────
    # Trust TMDB's own search ranking when the top result is clearly ahead of
    # everything else (margin >= 25 pts) but below the strict threshold.
    if top_score >= FALLBACK_THRESHOLD and top_score - second_score >= 25:
        return top_result, top_score

    return None, 0


# ── Result normalisation ───────────────────────────────────────────────────────

def normalise(result: dict, media_type: str) -> dict:
    """Extract the fields we care about from a raw TMDB result dict."""
    title_key = "title"          if media_type == "Movie" else "name"
    orig_key  = "original_title" if media_type == "Movie" else "original_name"
    date_key  = "release_date"   if media_type == "Movie" else "first_air_date"
    return {
        "matched":         True,
        "tmdb_id":         result["id"],
        "tmdb_title":      result.get(title_key, ""),
        "original_title":  result.get(orig_key, ""),
        "year":            (result.get(date_key, "") or "")[:4],
        "genre_ids":       result.get("genre_ids", []),
        "vote_average":    result.get("vote_average", 0),
        "media_type":      media_type,
    }


# ── Conflict log ───────────────────────────────────────────────────────────────

def load_conflicts() -> list:
    if CONFLICTS_FILE.exists():
        return json.loads(CONFLICTS_FILE.read_text(encoding="utf-8"))
    return []


def save_conflicts(conflicts: list) -> None:
    save_json(CONFLICTS_FILE, conflicts)


# ── watch_history.xlsx helpers ─────────────────────────────────────────────────

def add_rating_column() -> None:
    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if "Your Rating" in headers:
        return
    cell = ws.cell(1, ws.max_column + 1)
    cell.value = "Your Rating"
    cell.font = Font(bold=True)
    try:
        wb.save(WATCH_FILE)
        print(f"Added 'Your Rating' column to {WATCH_FILE.name}.")
        print("Open watch_history.xlsx and enter a rating (1-5) for titles you want to weight.")
    except PermissionError:
        print(f"WARNING: Could not add 'Your Rating' column — {WATCH_FILE.name} is open in Excel.")


# ── test / debug mode ─────────────────────────────────────────────────────────

def cmd_test(title: str, media_type: str, api_key: str) -> None:
    """Search TMDB for a single title and print all results with scores.

    Usage:  python recommender/01_match_tmdb.py --test "Title Here" [Movie|Series]
    """
    search_title = clean_for_search(title)
    print(f"Input:          {title!r}")
    if search_title != title:
        print(f"Searching as:   {search_title!r}  (platform suffix stripped)")
    print(f"Type:           {media_type}")
    print()
    results = search_tmdb(search_title, media_type, api_key)
    if not results:
        print("  No results returned by TMDB.")
        return

    title_key = "title"          if media_type == "Movie" else "name"
    orig_key  = "original_title" if media_type == "Movie" else "original_name"
    date_key  = "release_date"   if media_type == "Movie" else "first_air_date"

    print(f"  {'#':<3}  {'Score':>5}  {'tsr':>5}  {'pr':>5}  {'orig_tsr':>8}  {'Year':<6}  TMDB Title  (original_title)")
    print("  " + "-" * 90)
    for i, r in enumerate(results[:10], 1):
        tmdb_title = r.get(title_key, "") or ""
        orig_title = r.get(orig_key,  "") or ""
        tsr      = fuzz.token_sort_ratio(search_title.lower(), tmdb_title.lower())
        pr       = fuzz.partial_ratio(search_title.lower(),     tmdb_title.lower())
        orig_tsr = fuzz.token_sort_ratio(search_title.lower(), orig_title.lower()) if orig_title else 0
        score    = max(tsr, pr, orig_tsr)
        year     = (r.get(date_key, "") or "")[:4]
        flag     = " <-- would match" if score >= MATCH_THRESHOLD else (
                   " <-- fallback?"   if score >= FALLBACK_THRESHOLD else "")
        print(f"  {i:<3}  {score:>5}  {tsr:>5}  {pr:>5}  {orig_tsr:>8}  {year:<6}  {tmdb_title}  ({orig_title}){flag}")

    print()
    matched, score = best_match(search_title, results, media_type)
    if matched:
        print(f"  Result: MATCH -> {matched.get(title_key, '')} (score {score})")
    else:
        print("  Result: NO MATCH")


# ── main ───────────────────────────────────────────────────────────────────────

def main():
    reset   = "--reset" in sys.argv
    api_key = require_api_key()

    # --test "Title" [Movie|Series]
    if "--test" in sys.argv:
        idx        = sys.argv.index("--test")
        test_title = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else ""
        test_type  = sys.argv[idx + 2] if idx + 2 < len(sys.argv) else "Movie"
        if test_title:
            cmd_test(test_title, test_type, api_key)
        else:
            print("Usage: --test \"Title Here\" [Movie|Series]")
        return

    add_rating_column()

    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col  = {h: i for i, h in enumerate(headers)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # The 'url' column's cell *value* is the title name used as display text.
    # The actual Netflix/Amazon URL lives in the cell's hyperlink attribute.
    # Build a row-index → URL map by reading hyperlinks directly from the cells.
    url_col_1based = col.get("url", 3) + 1   # col dict is 0-based; cell() is 1-based
    row_urls: dict[int, str] = {}
    for r_idx in range(2, ws.max_row + 1):
        h = ws.cell(r_idx, url_col_1based).hyperlink
        if h:
            # openpyxl returns a Hyperlink object; .target holds the URL string
            row_urls[r_idx] = getattr(h, "target", "") or str(h)

    cache     = {} if reset else load_matches()
    conflicts = [] if reset else load_conflicts()
    # Keep a set of our_names already in the conflict log so we don't duplicate
    # entries on incremental re-runs.
    logged_names = {c["our_name"] for c in conflicts}

    unmatched  = []
    matched    = 0
    skipped    = 0
    confirmed  = 0   # fuzzy + netflix agreed
    overridden = 0   # netflix ID overrode fuzzy result
    total      = sum(1 for r in rows if r[col.get("Name", 1)])
    done       = 0

    for row_idx, row in enumerate(rows, start=2):
        name       = row[col.get("Name", 1)]
        media_type = row[col.get("Type", 0)]
        last_watch = row[col.get("last_watched", 4)]
        url        = row_urls.get(row_idx, "")
        if not name:
            continue

        done += 1

        if name in cache and not reset:
            skipped += 1
            continue

        watched_year = None
        if last_watch:
            try:
                watched_year = int(str(last_watch)[:4])
            except (ValueError, TypeError):
                pass

        netflix_id = extract_netflix_id(url)

        search_name = clean_for_search(name)
        suffix_note = f" [searching as: {search_name!r}]" if search_name != name else ""
        print(f"[{done}/{total}] {name} ({media_type}){suffix_note}...", end=" ", flush=True)
        try:
            # ── fuzzy search ──────────────────────────────────────────────────
            results      = search_tmdb(search_name, media_type, api_key)
            fuzzy_result, score = best_match(search_name, results, media_type, watched_year)
            time.sleep(0.25)

            # ── Netflix ID lookup ─────────────────────────────────────────────
            nf_hit = None
            if netflix_id:
                nf_hit = find_by_netflix_id(netflix_id, api_key)
                time.sleep(0.25)

            # ── reconcile ─────────────────────────────────────────────────────
            if fuzzy_result and nf_hit:
                fuzzy_id  = fuzzy_result["id"]
                nf_id     = nf_hit["result"]["id"]
                nf_mtype  = nf_hit["media_type"]
                nf_tkey   = "title" if nf_mtype == "Movie" else "name"
                nf_title  = nf_hit["result"].get(nf_tkey, "")
                fz_tkey   = "title" if media_type == "Movie" else "name"
                fz_title  = fuzzy_result.get(fz_tkey, "")

                if fuzzy_id == nf_id:
                    # Both agree — high confidence.
                    entry = normalise(fuzzy_result, media_type)
                    entry["match_score"]  = score
                    entry["match_method"] = "fuzzy+netflix_confirmed"
                    confirmed += 1
                    print(f"-> {fz_title} ({score}) [confirmed by Netflix ID]")
                else:
                    # Conflict — Netflix ID wins; log the disagreement.
                    entry = normalise(nf_hit["result"], nf_mtype)
                    entry["match_score"]  = None
                    entry["match_method"] = "netflix_id_override"
                    overridden += 1
                    print(f"-> {nf_title} [Netflix ID override; fuzzy had: {fz_title} ({score})]")

                    if name not in logged_names:
                        conflicts.append({
                            "our_name":          name,
                            "fuzzy_tmdb_title":  fz_title,
                            "fuzzy_tmdb_id":     fuzzy_id,
                            "fuzzy_score":       score,
                            "netflix_tmdb_title": nf_title,
                            "netflix_tmdb_id":   nf_id,
                            "resolved_to":       "netflix_id",
                        })
                        logged_names.add(name)
                        save_conflicts(conflicts)

            elif nf_hit:
                # Fuzzy failed but Netflix ID found something.
                nf_mtype = nf_hit["media_type"]
                nf_tkey  = "title" if nf_mtype == "Movie" else "name"
                nf_title = nf_hit["result"].get(nf_tkey, "")
                entry = normalise(nf_hit["result"], nf_mtype)
                entry["match_score"]  = None
                entry["match_method"] = "netflix_id_only"
                matched += 1
                print(f"-> {nf_title} [fuzzy miss, found via Netflix ID]")

                if name not in logged_names:
                    conflicts.append({
                        "our_name":          name,
                        "fuzzy_tmdb_title":  None,
                        "fuzzy_tmdb_id":     None,
                        "fuzzy_score":       0,
                        "netflix_tmdb_title": nf_title,
                        "netflix_tmdb_id":   nf_hit["result"]["id"],
                        "resolved_to":       "netflix_id",
                    })
                    logged_names.add(name)
                    save_conflicts(conflicts)

            elif fuzzy_result:
                # No Netflix ID (Amazon title) or Netflix lookup returned nothing.
                fz_tkey = "title" if media_type == "Movie" else "name"
                entry = normalise(fuzzy_result, media_type)
                entry["match_score"]  = score
                entry["match_method"] = "fuzzy_only"
                flag = " [fallback]" if score < MATCH_THRESHOLD else ""
                print(f"-> {fuzzy_result.get(fz_tkey, '')} ({score}){flag}")

            else:
                cache[name] = {"matched": False, "media_type": media_type}
                unmatched.append(name)
                print("-> NO MATCH")
                save_json(MATCHES_FILE, cache)
                continue

            cache[name] = entry
            matched += 1

        except Exception as e:
            print(f"-> ERROR: {e}")
            cache[name] = {"matched": False, "media_type": media_type, "error": str(e)}
            unmatched.append(name)

        save_json(MATCHES_FILE, cache)

    print()
    print(f"Done.  Matched: {matched}  |  Cached/skipped: {skipped}  |  Unmatched: {len(unmatched)}")
    if confirmed:
        print(f"  {confirmed} Netflix titles confirmed (fuzzy + Netflix ID agreed)")
    if overridden:
        print(f"  {overridden} titles overridden by Netflix ID (see {CONFLICTS_FILE.name})")
    if unmatched:
        print("  Unmatched titles will appear in the 'Unmatched' sheet of recommendations.xlsx")


if __name__ == "__main__":
    main()
