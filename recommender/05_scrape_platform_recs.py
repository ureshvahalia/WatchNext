"""
Step 5 -- Scrape recommendation carousels from Netflix and Prime Video.

For each watched title that has a Netflix or Amazon URL in watch_history.xlsx,
navigates to that title's page and extracts the "More Like This" /
"Customers also watched" recommendations using the already-saved login sessions.

Netflix strategy:
  Navigate to /title/{id}, look for a "More Like This" tab and click it,
  then extract all /title/ links on the page.

Amazon strategy:
  Navigate to the detail page, scroll down to load lazy carousels,
  then extract all /detail/ and /dp/ links.

Both fall back to harvesting all visible platform links on the page if the
specific recommendation section can't be found.

Cache: recommender/cache/platform_recs.json  (keyed by source URL)
Re-running skips already-scraped URLs unless --reset is passed.

Usage:
  python recommender/05_scrape_platform_recs.py
  python recommender/05_scrape_platform_recs.py --platform netflix
  python recommender/05_scrape_platform_recs.py --platform amazon
  python recommender/05_scrape_platform_recs.py --reset
  python recommender/05_scrape_platform_recs.py --limit 10   (test with first N)
  python recommender/05_scrape_platform_recs.py --debug      (save screenshots)
"""
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError
from thefuzz import fuzz

sys.path.insert(0, str(Path(__file__).parent))
from _common import WATCH_FILE, save_json

PROJECT_DIR   = Path(__file__).parent.parent
CACHE_FILE    = Path(__file__).parent / "cache" / "platform_recs.json"
DEBUG_DIR     = PROJECT_DIR / "output" / "debug"
NF_SESSION    = PROJECT_DIR / ".session" / "netflix"
AMZ_SESSION   = PROJECT_DIR / ".session"

PAGE_TIMEOUT    = 20_000   # ms to wait for page load
NAV_DELAY       = 1_000    # ms minimum pause after navigation (polite floor)
ELEMENT_TIMEOUT = 6_000    # ms to wait for a specific element to appear
SCROLL_STEPS    = 3        # number of scroll steps on Amazon pages

# ── JavaScript extractors ──────────────────────────────────────────────────────

_JS_NETFLIX = """
() => {
    const seen  = new Set();
    const items = [];

    // Netflix renders recommendation cards as role="button" divs, not <a> tags.
    // The title is in aria-label; the Netflix video ID is in the JSON stored in
    // data-ui-tracking-context on an inner element.
    // We scope to .moreLikeThis--wrapper so we only capture that section.
    const wrapper = document.querySelector('.moreLikeThis--wrapper');
    if (!wrapper) return items;   // section not present on this page

    const cards = wrapper.querySelectorAll('[data-uia="titleCard--container"]');
    for (const card of cards) {
        const title = (card.getAttribute('aria-label') || '').trim();
        if (!title || title.length < 2) continue;

        // The tracking context JSON is on the first child element that has it
        const ctxEl = card.querySelector('[data-ui-tracking-context]');
        if (!ctxEl) continue;

        let videoId = null;
        try {
            const ctx = JSON.parse(decodeURIComponent(ctxEl.getAttribute('data-ui-tracking-context')));
            videoId = ctx.video_id;
        } catch (e) { continue; }

        if (!videoId) continue;

        const url = 'https://www.netflix.com/title/' + videoId;
        if (seen.has(url)) continue;
        seen.add(url);

        items.push({ url, title });
    }
    return items;
}
"""

_JS_AMAZON = """
() => {
    const seen  = new Set();
    const items = [];

    // Capture the current page's detail ID to exclude it from results
    const currentHref = window.location.href.replace(/[?#].*$/, '').replace(/\\/ref=.*$/, '');

    const anchors = document.querySelectorAll(
        'a[href*="/detail/"], a[href*="/dp/"], a[href*="/gp/video/detail"]'
    );
    for (const a of anchors) {
        const href = a.href || '';
        if (!href) continue;

        const url = href.replace(/[?#].*$/, '').replace(/\\/ref=.*$/, '');
        if (url === currentHref) continue;
        if (seen.has(url)) continue;
        seen.add(url);

        const title = (a.textContent || a.getAttribute('aria-label') || a.getAttribute('title') || '').trim();
        if (!title || title.length < 2) continue;

        items.push({ url, title });
    }
    return items;
}
"""

# ── cache helpers ──────────────────────────────────────────────────────────────

def load_cache() -> dict:
    if CACHE_FILE.exists():
        return json.loads(CACHE_FILE.read_text(encoding="utf-8"))
    return {}


# ── watch history helpers ──────────────────────────────────────────────────────

def load_urls_from_watch_history(platform_filter: str | None) -> list[dict]:
    """Return list of {name, platform, url} for titles with a known platform URL."""
    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers)}

    url_col_1based = col.get("url", 3) + 1
    name_idx   = col.get("Name", 1)
    source_idx = col.get("source", 2)

    entries = []
    for r_idx in range(2, ws.max_row + 1):
        h = ws.cell(r_idx, url_col_1based).hyperlink
        if not h:
            continue
        url = getattr(h, "target", "") or str(h)
        if not url:
            continue

        row    = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
        name   = row[name_idx]   or ""
        source = row[source_idx] or ""

        if "netflix.com" in url:
            platform = "Netflix"
        elif "amazon.com" in url or "primevideo.com" in url:
            platform = "Amazon"
        else:
            continue

        if platform_filter and platform.lower() != platform_filter.lower():
            continue

        entries.append({"name": name, "platform": platform, "url": url})

    return entries


# ── Netflix scraping ───────────────────────────────────────────────────────────

async def scrape_netflix_page(page, url: str, debug: bool, debug_dir: Path) -> list[dict]:
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        await page.wait_for_timeout(NAV_DELAY)
    except PlaywrightTimeoutError:
        print("    TIMEOUT loading page")
        return []

    # Detect session expiry
    if "login" in page.url.lower():
        print("    SESSION EXPIRED — re-login required (run run.bat Netflix --fresh)")
        return []

    # Wait for the More Like This section specifically rather than sleeping blindly.
    # Falls back to the minimum NAV_DELAY if it never appears (title not on Netflix).
    try:
        await page.wait_for_selector(".moreLikeThis--wrapper", timeout=ELEMENT_TIMEOUT)
    except Exception:
        await page.wait_for_timeout(NAV_DELAY)

    if debug:
        debug_dir.mkdir(parents=True, exist_ok=True)
        slug = url.rstrip("/").split("/")[-1]
        await page.screenshot(path=str(debug_dir / f"nf_{slug}.png"))

    items = await page.evaluate(_JS_NETFLIX)
    return items


# ── Amazon scraping ────────────────────────────────────────────────────────────

async def scrape_amazon_page(page, url: str, debug: bool, debug_dir: Path) -> list[dict]:
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        await page.wait_for_timeout(NAV_DELAY)
    except PlaywrightTimeoutError:
        print("    TIMEOUT loading page")
        return []

    # Detect session expiry / sign-in redirect
    if "signin" in page.url.lower() or "ap/signin" in page.url.lower():
        print("    SESSION EXPIRED — re-login required (run run.bat Prime --fresh)")
        return []

    # Scroll to trigger lazy-loaded carousels, then wait for links to appear.
    for _ in range(SCROLL_STEPS):
        await page.keyboard.press("End")
        await page.wait_for_timeout(600)
    try:
        await page.wait_for_selector('a[href*="/detail/"], a[href*="/dp/"]',
                                      timeout=ELEMENT_TIMEOUT)
    except Exception:
        pass

    if debug:
        debug_dir.mkdir(parents=True, exist_ok=True)
        slug = url.rstrip("/").split("/")[-1][:20]
        await page.screenshot(path=str(debug_dir / f"amz_{slug}.png"))

    items = await page.evaluate(_JS_AMAZON)
    return items


# ── Amazon fallback (search-based) ────────────────────────────────────────────

AMZ_SEARCH_URL = "https://www.primevideo.com/search/ref=atv_nb_sr?phrase={query}"
SEARCH_MATCH_THRESHOLD = 75   # min fuzzy score to accept a search result as the right title


async def search_and_scrape_amazon(page, title: str,
                                    debug: bool, debug_dir: Path) -> tuple[list[dict], str]:
    """Search Prime Video for title, navigate to best match, return (recs, detail_url)."""
    search_url = AMZ_SEARCH_URL.format(query=quote(title))
    try:
        await page.goto(search_url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
        await page.wait_for_timeout(NAV_DELAY)
    except PlaywrightTimeoutError:
        return [], ""

    if "signin" in page.url.lower() or "ap/signin" in page.url.lower():
        return [], ""

    # Extract the first few detail-page links from the search results
    results = await page.evaluate(_JS_AMAZON)
    if not results:
        return [], ""

    # Pick the result whose title best matches ours
    best_url, best_score = "", 0
    for r in results[:5]:
        score = fuzz.token_sort_ratio(title.lower(), r["title"].lower())
        if score > best_score:
            best_score, best_url = score, r["url"]

    if best_score < SEARCH_MATCH_THRESHOLD or not best_url:
        return [], ""

    # Navigate to the detail page and scrape its recommendation carousel
    recs = await scrape_amazon_page(page, best_url, debug, debug_dir)
    return recs, best_url


async def scrape_amazon_fallbacks(fallback_entries: list[dict], cache: dict,
                                   debug: bool, debug_dir: Path) -> int:
    """For Netflix titles with 0 recs, search Prime Video and scrape there instead.

    Recs are stored under the original Netflix URL so re-runs skip these entries.
    """
    if not fallback_entries:
        return 0

    print(f"\n{'='*60}")
    print(f" Amazon fallback: searching Prime Video for "
          f"{len(fallback_entries)} Netflix titles with 0 recs")
    print(f"{'='*60}")

    if not AMZ_SESSION.exists():
        print("  No Amazon session found — skipping fallback.")
        return 0

    scraped = 0
    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=str(AMZ_SESSION),
            headless=False,
            viewport={"width": 1280, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        for i, entry in enumerate(fallback_entries, 1):
            nf_url = entry["url"]
            name   = entry["name"]
            print(f"  [{i}/{len(fallback_entries)}] {name}...", end=" ", flush=True)

            recs, detail_url = await search_and_scrape_amazon(page, name, debug, debug_dir)

            # Store under the Netflix URL so re-runs skip this entry
            cache[nf_url]["recs"]              = recs
            cache[nf_url]["fallback_platform"] = "Amazon"
            cache[nf_url]["fallback_url"]      = detail_url
            cache[nf_url]["fallback_tried"]    = True
            save_json(CACHE_FILE, cache)
            scraped += 1

            if recs:
                print(f"{len(recs)} recs  (via Prime Video)")
            elif detail_url:
                print("0 recs  (found on Prime but no carousel)")
            else:
                print("not found on Prime Video")

        await context.close()

    return scraped


# ── platform runner ────────────────────────────────────────────────────────────

_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)


async def scrape_platform(entries: list[dict], platform: str,
                           session_dir: Path, scrape_fn,
                           cache: dict, debug: bool, debug_dir: Path) -> int:
    platform_entries = [e for e in entries if e["platform"] == platform]
    todo = [e for e in platform_entries if e["url"] not in cache]

    if not platform_entries:
        return 0
    print(f"\n{'='*60}")
    print(f" {platform}: {len(todo)} to scrape, "
          f"{len(platform_entries)-len(todo)} already cached")
    print(f"{'='*60}")

    if not todo:
        return 0

    if not session_dir.exists():
        print(f"  No saved session found at {session_dir}.")
        print(f"  Run  run.bat {'Netflix' if platform == 'Netflix' else 'Prime'}  first.")
        return 0

    scraped = 0
    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=str(session_dir),
            headless=False,
            viewport={"width": 1280, "height": 900},
            user_agent=_USER_AGENT,
        )
        page = await context.new_page()

        for i, entry in enumerate(todo, 1):
            print(f"  [{i}/{len(todo)}] {entry['name']}...", end=" ", flush=True)
            try:
                items = await scrape_fn(page, entry["url"], debug, debug_dir)
            except Exception as exc:
                print(f"WARN: {type(exc).__name__}")
                items = []
            cache[entry["url"]] = {
                "source_title": entry["name"],
                "platform":     platform,
                "recs":         items,
                "scraped_at":   datetime.now().strftime("%Y-%m-%d"),
            }
            save_json(CACHE_FILE, cache)
            scraped += 1
            print(f"{len(items)} recs")

        await context.close()

    return scraped


# ── main ───────────────────────────────────────────────────────────────────────

def parse_args():
    platform = None
    reset    = "--reset" in sys.argv
    debug    = "--debug" in sys.argv
    limit    = None
    args     = sys.argv[1:]
    for i, a in enumerate(args):
        if a == "--platform" and i + 1 < len(args):
            platform = args[i + 1].capitalize()
        if a == "--limit" and i + 1 < len(args):
            try:
                limit = int(args[i + 1])
            except ValueError:
                pass
    return platform, reset, debug, limit


async def scrape_amazon_fallbacks(fallback_entries: list[dict], cache: dict,
                                   debug: bool, debug_dir: Path) -> int:
    """For Netflix titles with 0 recs, search Prime Video and scrape there instead."""
    if not fallback_entries:
        return 0

    print(f"\n{'='*60}")
    print(f" Amazon fallback: {len(fallback_entries)} Netflix titles with 0 recs")
    print(f"{'='*60}")

    if not AMZ_SESSION.exists():
        print("  No Amazon session found — skipping fallback.")
        return 0

    scraped = 0
    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=str(AMZ_SESSION),
            headless=False,
            viewport={"width": 1280, "height": 900},
            user_agent=_USER_AGENT,
        )
        page = await context.new_page()

        for i, entry in enumerate(fallback_entries, 1):
            nf_url = entry["url"]
            print(f"  [{i}/{len(fallback_entries)}] {entry['name']}...", end=" ", flush=True)
            try:
                recs, detail_url = await search_and_scrape_amazon(
                    page, entry["name"], debug, debug_dir
                )
            except Exception as exc:
                print(f"WARN: {type(exc).__name__}")
                recs, detail_url = [], ""

            cache[nf_url]["recs"]              = recs
            cache[nf_url]["fallback_platform"] = "Amazon"
            cache[nf_url]["fallback_url"]      = detail_url
            cache[nf_url]["fallback_tried"]    = True
            save_json(CACHE_FILE, cache)
            scraped += 1
            if recs:
                print(f"{len(recs)} recs via Prime")
            elif detail_url:
                print("found on Prime, 0 recs")
            else:
                print("not found on Prime")

        await context.close()

    return scraped


async def main():
    platform_filter, reset, debug, limit = parse_args()
    debug_dir = DEBUG_DIR / "platform_recs"

    entries = load_urls_from_watch_history(platform_filter)
    if limit:
        entries = entries[:limit]

    cache = {} if reset else load_cache()

    nf_scraped  = await scrape_platform(entries, "Netflix", NF_SESSION,
                                         scrape_netflix_page, cache, debug, debug_dir)
    amz_scraped = await scrape_platform(entries, "Amazon",  AMZ_SESSION,
                                         scrape_amazon_page,  cache, debug, debug_dir)

    fb_scraped = 0
    if platform_filter in (None, "Netflix"):
        netflix_entries  = [e for e in entries if e["platform"] == "Netflix"]
        fallback_entries = [
            e for e in netflix_entries
            if not cache.get(e["url"], {}).get("fallback_tried")
            and len(cache.get(e["url"], {}).get("recs", [])) == 0
        ]
        fb_scraped = await scrape_amazon_fallbacks(fallback_entries, cache, debug, debug_dir)

    print()
    print(f"Done.  Netflix: {nf_scraped}  |  Amazon: {amz_scraped}  "
          f"|  Amazon fallbacks: {fb_scraped}  |  Total cached: {len(cache)}")


if __name__ == "__main__":
    asyncio.run(main())
