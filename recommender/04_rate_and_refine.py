"""
Step 4 -- Re-rank recommendations using ratings (Method G).

Scoring: Hybrid 80/20 × Quality^1.0
  Taste  = 80% rated-only exponential weights + 20% frequency (both normalised)
  Quality = IMDB×0.6 + RT×0.4  (from quality_cache.json, built by compare_scoring.py)
            Falls back to TMDB vote_average ÷ 10 when quality cache is absent.
  Score   = taste × quality

Rated-only weights:  5★ → 5.2×   4★ → 2.3×   3★ → 1.0×   2★ → 0.43×   1★ → 0.19×
Unrated titles contribute only to the frequency (20%) component.
Falls back to pure frequency if fewer than 15 titles are rated.

Rewrites the Recommendations sheet in output/recommendations.xlsx.
The Unmatched sheet from step 3 is preserved.

Usage:
  python recommender/04_rate_and_refine.py
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent))
from _common import (
    GENRE_MAP, OUT_FILE, TMDB_BASE_URL, WATCH_FILE,
    load_matches, load_quality_cache, load_recs,
    score_method_g,
)


def load_ratings() -> dict:
    wb      = openpyxl.load_workbook(WATCH_FILE)
    ws      = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col     = {h: i for i, h in enumerate(headers)}
    name_idx   = col.get("Name", 1)
    rating_idx = col.get("Your Rating")
    if rating_idx is None:
        print("No 'Your Rating' column found — run Match first to add it.")
        return {}
    ratings: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        r    = row[rating_idx] if rating_idx < len(row) else None
        if name and r is not None:
            try:
                val = float(r)
                if 1.0 <= val <= 5.0:
                    ratings[str(name)] = val
            except (ValueError, TypeError):
                pass
    return ratings


def main():
    matches       = load_matches()
    recs          = load_recs()
    ratings       = load_ratings()
    quality_cache = load_quality_cache()

    if not ratings:
        print("No valid ratings found (1-5 in 'Your Rating'). Nothing to refine.")
        sys.exit(0)

    watched_ids = {str(m["tmdb_id"]) for m in matches.values() if m.get("matched")}

    q_source = "IMDB+RT" if quality_cache else "TMDB proxy (run Compare to fetch IMDB+RT)"
    print(f"Scoring with Method G  |  {len(ratings)} ratings  |  quality: {q_source}")

    ranked = score_method_g(matches, recs, ratings, watched_ids, quality_cache)
    print(f"Ranked {len(ranked):,} recommendations.")

    # ── Write Recommendations sheet ───────────────────────────────────────────
    wb = openpyxl.load_workbook(OUT_FILE) if OUT_FILE.exists() else openpyxl.Workbook()
    if "Recommendations" in wb.sheetnames:
        del wb["Recommendations"]
    ws = wb.create_sheet("Recommendations", 0)

    headers = ["Rank", "Title", "Type", "Freq", "Taste", "Quality",
               "Score", "TMDB Rating", "IMDB Rating", "RT %",
               "Genres", "Year", "TMDB Link"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for rank, item in enumerate(ranked, 1):
        genres = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in item["genre_ids"])
        seg    = "movie" if item["type"] == "Movie" else "tv"
        link   = f"{TMDB_BASE_URL}/{seg}/{item['tmdb_id']}"
        ws.append([
            rank,
            item["title"],
            item["type"],
            item["freq"],
            item["taste"],
            item["quality"],
            item["score"],
            round(item["vote_average"], 1),
            item["imdb_rating"],
            item["rt_score"],
            genres,
            item["year"],
            link,
        ])

    try:
        wb.save(OUT_FILE)
    except PermissionError:
        print(f"ERROR: Close {OUT_FILE.name} in Excel and retry.")
        sys.exit(1)

    print(f"Saved → {OUT_FILE}")


if __name__ == "__main__":
    main()
