"""
Compare recommendation scoring methods side by side.

Computes scoring approaches against your current ratings and writes
output/scoring_comparison.xlsx — one row per candidate title (union of the
top-100 from each method) with rank and score columns for every method.

Taste methods (A-E):
  A  Current       linear weight = rating/3,  unrated = 3★ (weight 1.0)
  B  Exponential   exp weight,                unrated = 3★ (weight 1.0)
  C  Rated-only    exp weight,                unrated = skipped entirely
  D  Hybrid 80/20  0.8 × rated-only + 0.2 × frequency  (both normalised)
  E  Genre profile genre-affinity weight × frequency × TMDB quality

Quality-enhanced hybrid (F-H) — hybrid taste × quality^γ:
  F  Hybrid + Quality γ=0.5   gentle quality influence
  G  Hybrid + Quality γ=1.0   moderate quality influence  (recommended)
  H  Hybrid + Quality γ=2.0   strong quality influence

Quality signal: IMDB rating (via OMDb) + RT Tomatometer if OMDB_API_KEY is set,
otherwise falls back to TMDB vote_average (correlated at ~0.88).

OMDb key setup (free, 1000 req/day):  https://www.omdbapi.com/apikey.aspx
  Option A — env var:   set OMDB_API_KEY=your_key
  Option B — file:      add  OMDB_API_KEY=your_key  to recommender/.env

Usage:
  python recommender/compare_scoring.py
"""
import math
import os
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from _common import (
    GENRE_MAP, MATCHES_FILE, RECS_FILE, WATCH_FILE,
    get_api_key, load_matches, load_recs, tmdb_score_factor,
)

# ── Config ────────────────────────────────────────────────────────────────────
TOP_N          = 100    # titles per method included in the candidate union
EXP_SIGMA      = 1.2   # exponential weight steepness
HYBRID_ALPHA   = 0.8   # rated-only fraction in the hybrid
RATED_FALLBACK = 15    # min rated titles before rated-only mode kicks in
QUALITY_GAMMAS = [0.5, 1.0, 2.0]   # γ values for quality-enhanced hybrid
IMDB_WEIGHT    = 0.6   # weight for IMDB in combined quality score
RT_WEIGHT      = 0.4   # weight for RT Tomatometer

_ROOT     = Path(__file__).parent.parent
OUT_FILE  = _ROOT / "output" / "scoring_comparison.xlsx"
QUAL_CACHE = _ROOT / "recommender" / "cache" / "quality_cache.json"

TMDB_BASE = "https://api.themoviedb.org/3"
OMDB_BASE = "http://www.omdbapi.com/"


# ── API key helpers ───────────────────────────────────────────────────────────

def get_omdb_key() -> str:
    key = os.environ.get("OMDB_API_KEY", "").strip()
    if key:
        return key
    env_file = _ROOT / "recommender" / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            if line.strip().startswith("OMDB_API_KEY="):
                return line.split("=", 1)[1].strip()
    return ""


# ── Data loading ──────────────────────────────────────────────────────────────

def load_ratings() -> dict:
    wb      = openpyxl.load_workbook(WATCH_FILE)
    ws      = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col     = {h: i for i, h in enumerate(headers)}
    name_idx   = col.get("Name", 1)
    rating_idx = col.get("Your Rating")
    out: dict = {}
    if rating_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_idx]
            r    = row[rating_idx] if rating_idx < len(row) else None
            if name and r is not None:
                try:
                    val = float(r)
                    if 1.0 <= val <= 5.0:
                        out[str(name)] = val
                except (ValueError, TypeError):
                    pass
    return out


def build_watched_ids(matches: dict) -> set:
    return {str(m["tmdb_id"]) for m in matches.values() if m.get("matched")}


def build_candidates(matches: dict, recs: dict, watched_ids: set) -> dict:
    seen: dict = {}
    for m in matches.values():
        if not m.get("matched"):
            continue
        for rec in recs.get(str(m["tmdb_id"]), []):
            tid = str(rec["tmdb_id"])
            if tid not in seen and tid not in watched_ids:
                seen[tid] = rec
    return seen


# ── Quality data: OMDb fetch + cache ─────────────────────────────────────────

def _load_quality_cache() -> dict:
    import json
    if QUAL_CACHE.exists():
        try:
            return json.loads(QUAL_CACHE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_quality_cache(cache: dict) -> None:
    import json
    QUAL_CACHE.parent.mkdir(parents=True, exist_ok=True)
    QUAL_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2),
                          encoding="utf-8")


def _tmdb_imdb_id(tmdb_id: int, media_type: str, tmdb_key: str) -> str | None:
    """Fetch the IMDB ID for a TMDB title via the external_ids endpoint."""
    seg = "movie" if media_type == "Movie" else "tv"
    params  = {} if tmdb_key.startswith("eyJ") else {"api_key": tmdb_key}
    headers = {"Authorization": f"Bearer {tmdb_key}"} if tmdb_key.startswith("eyJ") else {}
    try:
        r = requests.get(f"{TMDB_BASE}/{seg}/{tmdb_id}/external_ids",
                         headers=headers, params=params, timeout=8)
        if r.ok:
            return r.json().get("imdb_id") or None
    except Exception:
        pass
    return None


def _omdb_quality(imdb_id: str, omdb_key: str) -> dict:
    """Fetch IMDB rating and RT Tomatometer from OMDb."""
    result = {"imdb_rating": None, "rt_score": None, "metacritic": None}
    try:
        r = requests.get(OMDB_BASE, params={"i": imdb_id, "apikey": omdb_key}, timeout=8)
        if not r.ok:
            return result
        d = r.json()
        if d.get("Response") == "False":
            return result
        # IMDB rating
        ir = d.get("imdbRating", "N/A")
        if ir not in ("N/A", "", None):
            result["imdb_rating"] = float(ir)
        # RT and Metacritic from Ratings list
        for entry in d.get("Ratings", []):
            src = entry.get("Source", "")
            val = entry.get("Value", "")
            if src == "Rotten Tomatoes" and val.endswith("%"):
                result["rt_score"] = int(val[:-1])
            elif src == "Metacritic" and "/" in val:
                result["metacritic"] = int(val.split("/")[0])
    except Exception:
        pass
    return result


def fetch_quality_data(top_tids: set, candidates: dict) -> dict:
    """
    Return {tid: {"imdb_rating": float|None, "rt_score": int|None,
                  "quality": float, "source": str}}

    quality is the blended score in [0, 1]:
      - With OMDb: IMDB_WEIGHT × imdb/10 + RT_WEIGHT × rt/100
      - Without OMDb (or missing data): tmdb_vote_avg / 10

    Results are cached in quality_cache.json so reruns are instant.
    """
    cache    = _load_quality_cache()
    omdb_key = get_omdb_key()
    tmdb_key = get_api_key()

    to_fetch = [tid for tid in top_tids if tid not in cache]

    if not to_fetch:
        print(f"  Quality data: {len(top_tids)} titles from cache")
    elif omdb_key and tmdb_key:
        print(f"  Quality data: fetching {len(to_fetch)} titles from TMDB + OMDb "
              f"({len(cache)} already cached)...")
        for i, tid in enumerate(to_fetch, 1):
            d      = candidates.get(tid, {})
            mtype  = d.get("type", "Movie")
            entry  = {"imdb_id": None, "imdb_rating": None,
                      "rt_score": None, "metacritic": None}

            imdb_id = _tmdb_imdb_id(int(tid), mtype, tmdb_key)
            time.sleep(0.15)
            if imdb_id:
                entry["imdb_id"] = imdb_id
                omdb = _omdb_quality(imdb_id, omdb_key)
                entry.update(omdb)
                time.sleep(0.15)

            cache[tid] = entry
            if i % 20 == 0:
                _save_quality_cache(cache)
                print(f"    {i}/{len(to_fetch)}...")

        _save_quality_cache(cache)
        print(f"  Done — quality cache now has {len(cache)} entries")
    else:
        if not omdb_key:
            print("  Quality data: no OMDB_API_KEY found — using TMDB vote_average as proxy")
            print("  (get a free key at https://www.omdbapi.com/apikey.aspx)")
        for tid in to_fetch:
            cache[tid] = {"imdb_id": None, "imdb_rating": None,
                          "rt_score": None, "metacritic": None}

    # Build final quality dict
    result = {}
    for tid in top_tids:
        entry      = cache.get(tid, {})
        imdb       = entry.get("imdb_rating")
        rt         = entry.get("rt_score")
        tmdb_avg   = (candidates.get(tid) or {}).get("vote_average") or 5.0

        if imdb is not None and rt is not None:
            quality = IMDB_WEIGHT * (imdb / 10.0) + RT_WEIGHT * (rt / 100.0)
            source  = f"IMDB {imdb} / RT {rt}%"
        elif imdb is not None:
            quality = imdb / 10.0
            source  = f"IMDB {imdb}"
        else:
            quality = tmdb_avg / 10.0
            source  = f"TMDB {tmdb_avg} (proxy)"

        result[tid] = {
            "imdb_rating": imdb,
            "rt_score":    rt,
            "metacritic":  entry.get("metacritic"),
            "quality":     round(quality, 3),
            "source":      source,
        }
    return result


# ── Generic scorer ────────────────────────────────────────────────────────────

def _score(matches: dict, recs: dict, candidates: dict, ratings: dict,
           weight_fn) -> dict:
    """
    weight_fn(name, rating_or_None) -> float | None  (None = skip source)
    Returns {tid: {"score": float, "wtd": float, "freq": int}}
      score = wtd × log(1 + vote_avg)   (current quality blending)
      wtd   = raw weighted sum (quality-free), used by quality-enhanced methods
    """
    wtd:  dict = defaultdict(float)
    freq: dict = defaultdict(int)

    for name, m in matches.items():
        if not m.get("matched"):
            continue
        w = weight_fn(name, ratings.get(name))
        if w is None:
            continue
        sid  = str(m["tmdb_id"])
        seen: set = set()
        for rec in recs.get(sid, []):
            tid = str(rec["tmdb_id"])
            if tid in seen or tid not in candidates:
                continue
            seen.add(tid)
            wtd[tid]  += w
            freq[tid] += 1

    return {
        tid: {
            "score": round(wtd[tid] * tmdb_score_factor(candidates[tid].get("vote_average")), 4),
            "wtd":   round(wtd[tid], 4),
            "freq":  freq[tid],
        }
        for tid in wtd
    }


# ── Taste methods A-E ─────────────────────────────────────────────────────────

def method_current(matches, recs, candidates, ratings):
    """A — Linear weights; unrated treated as 3★."""
    return _score(matches, recs, candidates, ratings,
                  lambda n, r: (r if r is not None else 3.0) / 3.0)


def method_exponential(matches, recs, candidates, ratings):
    """B — Exponential weights; unrated treated as 3★."""
    return _score(matches, recs, candidates, ratings,
                  lambda n, r: math.exp(((r if r is not None else 3.0) - 3.0) / EXP_SIGMA))


def method_rated_only(matches, recs, candidates, ratings):
    """C — Exponential weights; unrated skipped. Frequency fallback if too few rated."""
    if len(ratings) < RATED_FALLBACK:
        print(f"    [C] only {len(ratings)} rated titles — using frequency fallback")
        return _score(matches, recs, candidates, ratings, lambda n, r: 1.0)
    return _score(matches, recs, candidates, ratings,
                  lambda n, r: None if r is None else math.exp((r - 3.0) / EXP_SIGMA))


def _hybrid_raw(matches, recs, candidates, ratings) -> dict:
    """Hybrid weighted sums WITHOUT the log(1+vote_avg) quality factor.
    Used as the taste signal for quality-enhanced methods F-H."""
    rated = method_rated_only(matches, recs, candidates, ratings)
    freq  = _score(matches, recs, candidates, ratings, lambda n, r: 1.0)

    max_r = max((v["wtd"] for v in rated.values()), default=1) or 1
    max_f = max((v["wtd"] for v in freq.values()),  default=1) or 1

    result: dict = {}
    for tid in set(rated) | set(freq):
        r_norm = rated.get(tid, {}).get("wtd", 0) / max_r
        f_norm = freq.get(tid,  {}).get("wtd", 0) / max_f
        result[tid] = {
            "taste": round(HYBRID_ALPHA * r_norm + (1 - HYBRID_ALPHA) * f_norm, 4),
            "freq":  freq.get(tid, {}).get("freq", 0),
        }
    return result


def method_hybrid(matches, recs, candidates, ratings):
    """D — HYBRID_ALPHA × rated-only + (1-HYBRID_ALPHA) × frequency (score-normalised)."""
    rated = method_rated_only(matches, recs, candidates, ratings)
    freq  = _score(matches, recs, candidates, ratings, lambda n, r: 1.0)

    max_r = max((v["score"] for v in rated.values()), default=1) or 1
    max_f = max((v["score"] for v in freq.values()),  default=1) or 1

    result: dict = {}
    for tid in set(rated) | set(freq):
        r_norm = rated.get(tid, {}).get("score", 0) / max_r
        f_norm = freq.get(tid,  {}).get("score", 0) / max_f
        result[tid] = {
            "score": round(HYBRID_ALPHA * r_norm + (1 - HYBRID_ALPHA) * f_norm, 4),
            "freq":  freq.get(tid, {}).get("freq", 0),
        }
    return result


def method_genre_profile(matches, recs, candidates, ratings):
    """E — Genre affinity from rated titles × frequency × TMDB quality."""
    g_sum:   dict = defaultdict(float)
    g_count: dict = defaultdict(int)
    for name, m in matches.items():
        if not m.get("matched"):
            continue
        r = ratings.get(name)
        if r is None:
            continue
        for gid in m.get("genre_ids", []):
            g_sum[gid]   += r
            g_count[gid] += 1
    profile = {gid: g_sum[gid] / g_count[gid] for gid in g_sum}

    freq: dict = defaultdict(int)
    for m in matches.values():
        if not m.get("matched"):
            continue
        seen: set = set()
        for rec in recs.get(str(m["tmdb_id"]), []):
            tid = str(rec["tmdb_id"])
            if tid in seen or tid not in candidates:
                continue
            seen.add(tid)
            freq[tid] += 1

    result: dict = {}
    for tid, f in freq.items():
        d      = candidates[tid]
        genres = d.get("genre_ids", [])
        affinity = (sum(profile.get(g, 3.0) for g in genres) / len(genres)
                    if genres and profile else 3.0)
        result[tid] = {
            "score":    round((affinity / 3.0) * f * tmdb_score_factor(d.get("vote_average")), 4),
            "affinity": round(affinity, 2),
            "freq":     f,
        }
    return result


# ── Quality-enhanced hybrid methods F-H ──────────────────────────────────────

def method_hybrid_quality(hybrid_raw: dict, quality_data: dict, gamma: float) -> dict:
    """
    F/G/H — Hybrid taste signal × quality^γ.

    taste is the normalised hybrid weighted sum (no quality baked in).
    quality is the blended IMDB+RT score (or TMDB proxy), in [0, 1].
    final_score = taste × quality^γ
    """
    result: dict = {}
    for tid, v in hybrid_raw.items():
        q     = quality_data.get(tid, {}).get("quality", 0.5)
        score = v["taste"] * (q ** gamma)
        result[tid] = {
            "score": round(score, 4),
            "freq":  v["freq"],
        }
    return result


# ── Ranking ───────────────────────────────────────────────────────────────────

def make_ranks(scores: dict) -> dict:
    ordered = sorted(scores, key=lambda t: -scores[t]["score"])
    return {tid: i + 1 for i, tid in enumerate(ordered)}


# ── Excel ─────────────────────────────────────────────────────────────────────

TASTE_META = [
    ("A Current",      "DDEEFF"),
    ("B Exponential",  "FFE8CC"),
    ("C Rated-only",   "DDFFD8"),
    ("D Hybrid 80/20", "FFD8E8"),
    ("E Genre profile","EAE0FF"),
]
QUALITY_META = [
    ("F Hybrid+Q γ=0.5", "FFF0C0"),
    ("G Hybrid+Q γ=1.0", "FFE080"),
    ("H Hybrid+Q γ=2.0", "FFB830"),
]
ALL_META = TASTE_META + QUALITY_META


def write_excel(sorted_tids, candidates, methods, ranks, ratings, quality_data):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Quality-Enhanced Focus ──────────────────────────────────────
    ws = wb.active
    ws.title = "Quality-Enhanced"

    q_keys  = [m[0] for m in QUALITY_META]
    d_key   = "D Hybrid 80/20"
    focus_keys = [d_key] + q_keys

    base    = ["Title", "Year", "Type", "TMDB ★", "IMDB ★", "RT %", "Quality",
               "Genres", "Freq", "Avg Rank (F-H)"]
    headers = base[:]
    for label in focus_keys:
        headers += [f"{label} Rank", f"{label} Score"]

    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Colour bands: D = pink, F/G/H = yellow shades
    base_n = len(base)
    col_meta = [(d_key, "FFD8E8")] + [(m[0], m[1]) for m in QUALITY_META]
    for mi, (label, colour) in enumerate(col_meta):
        fill = PatternFill("solid", fgColor=colour)
        for ci in [base_n + 1 + mi * 2, base_n + 2 + mi * 2]:
            ws.cell(1, ci).fill = fill

    ws.row_dimensions[1].height = 34

    for tid in sorted_tids:
        d      = candidates[tid]
        genres = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in d.get("genre_ids", []))
        freq   = methods[d_key].get(tid, {}).get("freq", 0)
        qd     = quality_data.get(tid, {})
        avg_rk = round(
            sum(ranks[k].get(tid, len(candidates) + 1) for k in q_keys) / len(q_keys), 1
        )
        row = [
            d["title"], d.get("year", ""), d.get("type", ""),
            round(d.get("vote_average") or 0, 1),
            qd.get("imdb_rating", "—"),
            qd.get("rt_score", "—"),
            qd.get("quality", "—"),
            genres, freq, avg_rk,
        ]
        for key in focus_keys:
            rk    = ranks[key].get(tid, "—")
            score = methods[key].get(tid, {}).get("score", "—")
            row  += [rk, score]
        ws.append(row)

    ws.column_dimensions["A"].width = 36
    for col, w in zip("BCDEFGHIJ", [6, 9, 8, 8, 6, 8, 30, 6, 11]):
        ws.column_dimensions[col].width = w
    for ci in range(base_n + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Full Comparison (all A-H) ───────────────────────────────────
    ws2     = wb.create_sheet("All Methods")
    all_keys = [m[0] for m in ALL_META]
    base2   = ["Title", "Year", "Type", "TMDB ★", "Quality", "Genres", "Freq", "Avg Rank"]
    hdrs2   = base2[:]
    for label in all_keys:
        hdrs2 += [f"{label} Rank", f"{label} Score"]
    ws2.append(hdrs2)
    for cell in ws2[1]:
        cell.font      = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    base2_n = len(base2)
    for mi, (label, colour) in enumerate(ALL_META):
        fill = PatternFill("solid", fgColor=colour)
        for ci in [base2_n + 1 + mi * 2, base2_n + 2 + mi * 2]:
            ws2.cell(1, ci).fill = fill
    ws2.row_dimensions[1].height = 28

    for tid in sorted_tids:
        d      = candidates[tid]
        genres = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in d.get("genre_ids", []))
        freq   = methods["A Current"].get(tid, {}).get("freq", 0)
        qd     = quality_data.get(tid, {})
        avg_rk = round(
            sum(ranks[k].get(tid, len(candidates) + 1) for k in all_keys) / len(all_keys), 1
        )
        row = [d["title"], d.get("year", ""), d.get("type", ""),
               round(d.get("vote_average") or 0, 1),
               qd.get("quality", "—"), genres, freq, avg_rk]
        for key in all_keys:
            rk    = ranks[key].get(tid, "—")
            score = methods[key].get(tid, {}).get("score", "—")
            row  += [rk, score]
        ws2.append(row)

    ws2.column_dimensions["A"].width = 36
    for col, w in zip("BCDEFGH", [6, 9, 8, 8, 30, 6, 11]):
        ws2.column_dimensions[col].width = w
    for ci in range(base2_n + 1, len(hdrs2) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 12
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # ── Sheet 3: Rated Titles ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Rated Titles")
    ws3.append(["Title", "Your Rating", "TMDB ID", "Genres"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    matches_ref = load_matches()
    for name, r in sorted(ratings.items(), key=lambda x: -x[1]):
        m      = matches_ref.get(name, {})
        genres = ", ".join(GENRE_MAP.get(gid, str(gid)) for gid in m.get("genre_ids", []))
        ws3.append([name, r, m.get("tmdb_id", ""), genres])
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 30

    # ── Sheet 4: Method Notes ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Method Notes")
    omdb_key = get_omdb_key()
    q_source = (f"IMDB ({IMDB_WEIGHT*100:.0f}%) + RT Tomatometer ({RT_WEIGHT*100:.0f}%)"
                if omdb_key else "TMDB vote_average ÷ 10  [OMDb key not set — proxy only]")
    notes = [
        ["Method", "Weight formula", "Unrated titles", "Quality signal", "Key characteristic"],
        ["A Current",       "rating / 3.0",                              "3★ (weight 1.0)", "log(1+TMDB★)",  "Baseline; all titles contribute equally when unrated"],
        ["B Exponential",   f"exp((r−3)/{EXP_SIGMA})",                  "3★ (weight 1.0)", "log(1+TMDB★)",  "Wider taste spread; 5★→5.2×, 1★→0.19×; still diluted"],
        ["C Rated-only",    f"exp((r−3)/{EXP_SIGMA})",                  "Skipped",         "log(1+TMDB★)",  "Pure taste signal; only your rated titles shape results"],
        ["D Hybrid 80/20",  f"{int(HYBRID_ALPHA*100)}% C + {int((1-HYBRID_ALPHA)*100)}% freq", "Skipped in rated half", "log(1+TMDB★)", "Balances personal taste with broad discovery"],
        ["E Genre profile", "genre-affinity × freq",                     "Excluded",        "log(1+TMDB★)",  "Generalises across genres; robust with sparse ratings"],
        ["F Hybrid+Q γ=0.5",f"D taste × quality^0.5",                   "Skipped in rated",""+q_source,     "Gentle quality boost; low-quality films lose ~10-20%"],
        ["G Hybrid+Q γ=1.0",f"D taste × quality^1.0",                   "Skipped in rated",""+q_source,     "Moderate quality boost; 8.5 vs 6.0 film → 42% spread"],
        ["H Hybrid+Q γ=2.0",f"D taste × quality^2.0",                   "Skipped in rated",""+q_source,     "Strong quality filter; low-quality films heavily penalised"],
        [],
        [f"Quality formula: {IMDB_WEIGHT} × (IMDB/10) + {RT_WEIGHT} × (RT/100)  "
         f"[or TMDB/10 when OMDb key unavailable]"],
        [f"Quality data source: {'OMDb (live IMDB + RT data)' if omdb_key else 'TMDB proxy — add OMDB_API_KEY for real IMDB + RT scores'}"],
    ]
    for row in notes:
        ws4.append(row)
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for col, w in zip("ABCDE", [20, 32, 22, 40, 52]):
        ws4.column_dimensions[col].width = w

    try:
        wb.save(OUT_FILE)
    except PermissionError:
        print(f"ERROR: Close {OUT_FILE.name} in Excel first, then retry.")
        sys.exit(1)
    print(f"Saved → {OUT_FILE}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    if not MATCHES_FILE.exists() or not RECS_FILE.exists():
        print("ERROR: Run phases 4+5 first (Match + Recommend) to build the caches.")
        sys.exit(1)

    matches     = load_matches()
    recs        = load_recs()
    ratings     = load_ratings()
    watched_ids = build_watched_ids(matches)
    candidates  = build_candidates(matches, recs, watched_ids)

    print(f"  {len(ratings)} rated title(s)  |  {len(candidates):,} recommendation candidates")

    if not ratings:
        print("\nNo ratings found — rate some titles in the UI (run.bat UI) first.")
        sys.exit(0)

    print("\nComputing taste methods (A-E)...")
    hybrid_raw = _hybrid_raw(matches, recs, candidates, ratings)

    methods: dict = {}
    for label, fn in [
        ("A Current",      lambda: method_current(matches, recs, candidates, ratings)),
        ("B Exponential",  lambda: method_exponential(matches, recs, candidates, ratings)),
        ("C Rated-only",   lambda: method_rated_only(matches, recs, candidates, ratings)),
        ("D Hybrid 80/20", lambda: method_hybrid(matches, recs, candidates, ratings)),
        ("E Genre profile",lambda: method_genre_profile(matches, recs, candidates, ratings)),
    ]:
        print(f"  {label}...", end=" ", flush=True)
        methods[label] = fn()
        print(f"{len(methods[label]):,} candidates")

    # Union of top-N from taste methods for quality fetching
    top_tids: set = set()
    for key in [m[0] for m in TASTE_META]:
        top_tids |= {tid for tid, r in make_ranks(methods[key]).items() if r <= TOP_N}

    print(f"\nFetching quality data for {len(top_tids)} titles...")
    quality_data = fetch_quality_data(top_tids, candidates)

    print("\nComputing quality-enhanced methods (F-H)...")
    for gamma, (label, _) in zip(QUALITY_GAMMAS, QUALITY_META):
        print(f"  {label}...", end=" ", flush=True)
        methods[label] = method_hybrid_quality(hybrid_raw, quality_data, gamma)
        # Only include candidates that are in the top_tids set (we have quality data for them)
        methods[label] = {tid: v for tid, v in methods[label].items() if tid in top_tids}
        print(f"{len(methods[label]):,} candidates")

    ranks = {k: make_ranks(v) for k, v in methods.items()}

    # Final union of top-N across all methods
    all_keys = [m[0] for m in ALL_META]
    final_tids: set = set()
    for key in all_keys:
        final_tids |= {tid for tid, r in ranks[key].items() if r <= TOP_N}
    print(f"\n  {len(final_tids)} unique titles in top-{TOP_N} across all methods")

    # Sort by average rank across quality methods F-H (the focus of this comparison)
    q_keys = [m[0] for m in QUALITY_META]
    def avg_q_rank(tid):
        return sum(ranks[k].get(tid, len(candidates) + 1) for k in q_keys) / len(q_keys)
    sorted_tids = sorted(final_tids, key=avg_q_rank)

    print("\nRatings used:")
    for name, r in sorted(ratings.items(), key=lambda x: -x[1]):
        print(f"  {'★' * int(r):<5}  {name}")

    print()
    write_excel(sorted_tids, candidates, methods, ranks, ratings, quality_data)


if __name__ == "__main__":
    main()
