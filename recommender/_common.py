"""Shared utilities for the recommendation pipeline."""
import json
import math
import os
import sys
from pathlib import Path

import openpyxl
import requests
from thefuzz import fuzz

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT           = Path(os.environ['WATCHNEXT_HOME']) if 'WATCHNEXT_HOME' in os.environ else Path(__file__).parent.parent
WATCH_FILE     = ROOT / "output" / "watch_history.xlsx"
MATCHES_FILE   = ROOT / "recommender" / "cache" / "tmdb_matches.json"
RECS_FILE      = ROOT / "recommender" / "cache" / "recs_raw.json"
OUT_FILE       = ROOT / "output" / "recommendations.xlsx"

# ── Config ─────────────────────────────────────────────────────────────────────
FILTER_THRESHOLD = 90    # min fuzzy score to flag a candidate as already watched
TMDB_BASE_URL    = "https://www.themoviedb.org"

# ── TMDB genre ID → name (movie + TV combined) ────────────────────────────────
GENRE_MAP = {
    28: "Action",        12: "Adventure",     16: "Animation",   35: "Comedy",
    80: "Crime",         99: "Documentary",   18: "Drama",    10751: "Family",
    14: "Fantasy",       36: "History",       27: "Horror",   10402: "Music",
  9648: "Mystery",    10749: "Romance",      878: "Sci-Fi",  10770: "TV Movie",
    53: "Thriller",   10752: "War",           37: "Western",
 10759: "Action & Adventure", 10762: "Kids",   10763: "News",  10764: "Reality",
 10765: "Sci-Fi & Fantasy",  10766: "Soap",   10767: "Talk", 10768: "War & Politics",
}


def get_api_key() -> str:
    """Return TMDB API key from TMDB_API_KEY env var or recommender/.env file."""
    key = os.environ.get("TMDB_API_KEY", "").strip()
    if key:
        return key
    env_file = ROOT / "recommender" / ".env"
    if env_file.exists():
        for line in env_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("TMDB_API_KEY="):
                return line.split("=", 1)[1].strip()
    return ""


def require_api_key() -> str:
    """Return the API key or print setup instructions and exit."""
    key = get_api_key()
    if not key:
        print("ERROR: TMDB API key not found.")
        print()
        print("To get a free key:")
        print("  1. Create an account at  https://www.themoviedb.org/signup")
        print("  2. Go to                 https://www.themoviedb.org/settings/api")
        print("  3. Click 'Create' and select Developer / personal use")
        print("  4. Copy either value — both are accepted:")
        print("       API Key (v3 auth)         short alphanumeric string")
        print("       Read Access Token (v4)    long string starting with eyJ...")
        print()
        print("Then store your key in ONE of these ways:")
        print()
        print("  Option A — environment variable (current session only):")
        print("    set TMDB_API_KEY=your_key_here")
        print()
        print("  Option B — file (persists across sessions):")
        print(r"    Create the file  recommender\.env  containing:")
        print("    TMDB_API_KEY=your_key_here")
        sys.exit(1)
    return key


def tmdb_get(url: str, api_key: str, params: dict | None = None) -> requests.Response:
    """GET a TMDB endpoint, handling both v3 API key and v4 Read Access Token.

    TMDB issues two credential types:
      - API Key (v3): short alphanumeric, passed as ?api_key=...
      - Read Access Token (v4): JWT starting with eyJ, sent as Bearer header.
    Both authenticate v3 endpoints; the token just travels differently.
    """
    params = params or {}
    if api_key.startswith("eyJ"):
        r = requests.get(url, params=params,
                         headers={"Authorization": f"Bearer {api_key}"},
                         timeout=10)
    else:
        r = requests.get(url, params={"api_key": api_key, **params}, timeout=10)
    r.raise_for_status()
    return r


def load_watch_history() -> tuple[dict, list]:
    """Return (col_index_map, rows).

    col_index_map: {header_name: 0-based index into the values tuple}
    rows: list of raw value tuples (one per data row)
    """
    wb = openpyxl.load_workbook(WATCH_FILE)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i for i, h in enumerate(headers)}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return col, rows


def watched_title_set() -> set:
    """Return set of lowercased watched title names."""
    col, rows = load_watch_history()
    name_idx = col.get("Name", 1)
    return {str(row[name_idx]).lower().strip() for row in rows if row[name_idx]}


def is_already_watched(title: str, watched_set: set) -> bool:
    t = title.lower().strip()
    if t in watched_set:
        return True
    return any(fuzz.token_sort_ratio(t, wt) >= FILTER_THRESHOLD for wt in watched_set)


def load_matches() -> dict:
    if not MATCHES_FILE.exists():
        return {}
    return json.loads(MATCHES_FILE.read_text(encoding="utf-8"))


def load_recs() -> dict:
    if not RECS_FILE.exists():
        return {}
    return json.loads(RECS_FILE.read_text(encoding="utf-8"))


def save_json(path: Path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def tmdb_score_factor(vote_avg) -> float:
    """log(1 + v) multiplier for blending TMDB rating into a frequency score.

    Titles with no recorded rating (0 or None) use 5.0 as a neutral midpoint
    rather than being zeroed out — absent data != bad quality.
    """
    v = vote_avg if vote_avg else 5.0
    return math.log(1 + v)


# ── Method G: Hybrid 80/20 × Quality^1.0 ─────────────────────────────────────
# These constants are shared by 04_rate_and_refine.py and ui/server.py.

QUALITY_CACHE_FILE = ROOT / "recommender" / "cache" / "quality_cache.json"

EXP_SIGMA      = 1.2   # exponential taste-weight steepness
HYBRID_ALPHA   = 0.8   # fraction from rated-only vs frequency
RATED_FALLBACK = 15    # min rated titles before rated-only mode engages
QUALITY_GAMMA  = 1.0   # quality exponent (1.0 = linear)
IMDB_WEIGHT    = 0.6   # IMDB share of combined quality score
RT_WEIGHT      = 0.4   # RT Tomatometer share


def load_quality_cache() -> dict:
    if not QUALITY_CACHE_FILE.exists():
        return {}
    try:
        return json.loads(QUALITY_CACHE_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def get_quality(tid: str, vote_avg, quality_cache: dict) -> float:
    """Blended quality in [0, 1].

    Uses IMDB + RT from quality_cache when available; falls back to
    TMDB vote_average / 10 (correlated at ~0.88 with IMDB).
    """
    entry = quality_cache.get(str(tid), {})
    imdb  = entry.get("imdb_rating")
    rt    = entry.get("rt_score")
    if imdb is not None and rt is not None:
        return IMDB_WEIGHT * (imdb / 10.0) + RT_WEIGHT * (rt / 100.0)
    if imdb is not None:
        return imdb / 10.0
    v = vote_avg if vote_avg else 5.0
    return v / 10.0


def score_method_g(matches: dict, recs: dict, ratings: dict,
                   watched_ids: set, quality_cache: dict) -> list:
    """Method G — Hybrid 80/20 × Quality^1.0.

    Taste signal: 80% rated-only (exp weights) + 20% frequency, both normalised.
    Quality signal: IMDB×0.6 + RT×0.4 from quality_cache, else TMDB/10 proxy.
    Final score: taste × quality^QUALITY_GAMMA

    Uses exact TMDB-ID matching to exclude watched titles (fast).
    Falls back to equal-weight frequency when fewer than RATED_FALLBACK
    titles are rated.

    Returns a list of dicts sorted by score descending:
      tmdb_id, title, type, year, genre_ids, vote_average,
      freq, taste, quality, imdb_rating, rt_score, score
    """
    from collections import defaultdict

    use_rated = len(ratings) >= RATED_FALLBACK

    rated_wtd: dict = defaultdict(float)
    freq_all:  dict = defaultdict(int)
    cand_data: dict = {}

    for name, m in matches.items():
        if not m.get("matched"):
            continue
        sid    = str(m["tmdb_id"])
        rating = ratings.get(name)

        if use_rated:
            w = math.exp((rating - 3.0) / EXP_SIGMA) if rating is not None else None
        else:
            w = 1.0  # frequency fallback: all titles equal

        seen: set = set()
        for rec in recs.get(sid, []):
            tid = str(rec["tmdb_id"])
            if tid in seen or tid in watched_ids:
                continue
            seen.add(tid)
            if tid not in cand_data:
                cand_data[tid] = rec
            freq_all[tid] += 1
            if w is not None:
                rated_wtd[tid] += w

    if not cand_data:
        return []

    max_rated = max(rated_wtd.values(), default=1) or 1
    max_freq  = max(freq_all.values(),  default=1) or 1

    results = []
    for tid, d in cand_data.items():
        r_norm  = rated_wtd.get(tid, 0) / max_rated
        f_norm  = freq_all[tid] / max_freq
        taste   = HYBRID_ALPHA * r_norm + (1 - HYBRID_ALPHA) * f_norm
        va      = d.get("vote_average")
        quality = get_quality(tid, va, quality_cache)
        entry   = quality_cache.get(tid, {})
        results.append({
            "tmdb_id":     int(tid),
            "title":       d["title"],
            "type":        d.get("type", ""),
            "year":        d.get("year", ""),
            "genre_ids":   d.get("genre_ids", []),
            "vote_average": va or 0,
            "freq":        freq_all[tid],
            "taste":       round(taste, 4),
            "quality":     round(quality, 3),
            "imdb_rating": entry.get("imdb_rating"),
            "rt_score":    entry.get("rt_score"),
            "score":       round(taste * (quality ** QUALITY_GAMMA), 4),
        })

    results.sort(key=lambda x: -x["score"])
    return results
