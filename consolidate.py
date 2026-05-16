#!/usr/bin/env python3
"""
Watch History Consolidator

Merges output/amazon_clean.csv and output/netflix_clean.csv into a single
output/watch_history.xlsx.

Titles that appear on both platforms are detected via fuzzy name matching
(token_sort_ratio >= 90) and merged into one row.  Types must also match —
a movie on one platform and a series on the other are kept as separate rows.
The single URL written per row is taken from whichever platform has the more
recent watch date.
"""

import csv
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from thefuzz import fuzz


_DATE_FORMATS = ("%B %d, %Y", "%m/%d/%Y", "%Y-%m-%d")

HYPERLINK_FONT = Font(color="0563C1", underline="single")


def parse_date(s: str) -> date | str:
    s = s.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s


def _newer(a: date | str, b: date | str) -> date | str:
    """Return whichever of a/b is a date and more recent; fall back to a."""
    if isinstance(a, date) and isinstance(b, date):
        return a if a >= b else b
    return a if isinstance(a, date) else b


PROJECT_DIR     = Path(__file__).parent
OUTPUT_DIR      = PROJECT_DIR / "output"
FUZZY_THRESHOLD = 90   # 0-100; token_sort_ratio handles word-order differences


def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def consolidate():
    amazon_path  = OUTPUT_DIR / "amazon_clean.csv"
    netflix_path = OUTPUT_DIR / "netflix_clean.csv"

    amazon_rows  = load_csv(amazon_path)
    netflix_rows = load_csv(netflix_path)

    if not amazon_rows and not netflix_rows:
        print(
            "ERROR: Neither amazon_clean.csv nor netflix_clean.csv found in output/.\n"
            "       Run the cleaners first (run.bat Process)."
        )
        sys.exit(1)

    if not amazon_rows:
        print("  Note: no Amazon data found - using Netflix only.")
    if not netflix_rows:
        print("  Note: no Netflix data found - using Amazon only.")

    print(f"Consolidating: {len(amazon_rows)} Amazon + {len(netflix_rows)} Netflix entries...")

    # Each result row: Type, Name, source, url, url_label, last_watched
    # url_label is the platform name used as the hyperlink display text.
    result:       list[dict] = []
    amazon_names: list[str]  = []   # lowercased, parallel to result[]

    for row in amazon_rows:
        result.append({
            "Type":         row.get("Type", ""),
            "Name":         row.get("Name", ""),
            "source":       "Amazon",
            "url":          row.get("url", ""),
            "url_label":    "Amazon",
            "last_watched": parse_date(row.get("date_watched", "")),
        })
        amazon_names.append(row.get("Name", "").lower())

    # Merge Netflix entries, fuzzy-matching against the Amazon list.
    # Only merge when both Type and name match — same title as movie on one
    # platform and series on the other (e.g. The Lincoln Lawyer) stays separate.
    merged           = 0
    new_from_netflix = 0

    for nf in netflix_rows:
        nf_name = nf.get("Name", "").lower()
        nf_type = nf.get("Type", "")

        best_score = 0
        best_idx   = -1
        for i, am_name in enumerate(amazon_names):
            score = fuzz.token_sort_ratio(nf_name, am_name)
            if score > best_score:
                best_score = score
                best_idx   = i

        type_matches = best_idx >= 0 and result[best_idx]["Type"] == nf_type

        if best_score >= FUZZY_THRESHOLD and type_matches:
            # Title found on both platforms with the same type — pick URL from
            # whichever platform has the more recent watch date.
            am_date = result[best_idx]["last_watched"]
            nf_date = parse_date(nf.get("date", ""))
            most_recent = _newer(am_date, nf_date)

            if most_recent is nf_date or (isinstance(nf_date, date) and nf_date >= am_date):
                result[best_idx]["url"]       = nf.get("url", "")
                result[best_idx]["url_label"] = "Netflix"
            # else keep the Amazon url already set

            result[best_idx]["last_watched"] = most_recent
            result[best_idx]["source"]       = "Both"
            merged += 1
        else:
            result.append({
                "Type":         nf_type,
                "Name":         nf.get("Name", ""),
                "source":       "Netflix",
                "url":          nf.get("url", ""),
                "url_label":    "Netflix",
                "last_watched": parse_date(nf.get("date", "")),
            })
            new_from_netflix += 1

    result.sort(key=lambda r: r["Name"].lower())

    # ── write Excel ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Watch History"

    headers = ["Type", "Name", "source", "url", "last_watched"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    url_col_idx  = headers.index("url")  + 1  # 1-based
    date_col_idx = headers.index("last_watched") + 1

    for row in result:
        ws.append([
            row["Type"],
            row["Name"],
            row["source"],
            row["Name"] if row["url"] else "",
            row["last_watched"],
        ])
        data_row = ws.max_row

        # Hyperlink on the url cell
        if row["url"]:
            cell = ws.cell(row=data_row, column=url_col_idx)
            cell.hyperlink = row["url"]
            cell.font = HYPERLINK_FONT

        # Date format on the last_watched cell
        cell = ws.cell(row=data_row, column=date_col_idx)
        if isinstance(cell.value, date):
            cell.number_format = "YYYY-MM-DD"

    # Auto-size columns (cap at 40 chars; URLs are now short labels)
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    out_path = OUTPUT_DIR / "watch_history.xlsx"
    try:
        wb.save(out_path)
    except PermissionError:
        print(f"\n  ERROR: Cannot write {out_path.name} - the file is open in Excel.")
        print("  Please close it and run again.")
        sys.exit(1)

    total = len(result)
    print(f"  {merged} titles found on both platforms (merged)")
    print(f"  {new_from_netflix} titles from Netflix only")
    print(f"  {len(amazon_rows) - merged} titles from Amazon only")
    print(f"  {total} total unique titles -> {out_path}")


if __name__ == "__main__":
    consolidate()
